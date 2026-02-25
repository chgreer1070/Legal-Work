"""
Script to create UserMapping.xlsx with team member data.
The resulting file can be uploaded to OneDrive and formatted as a table named "UserMap".
"""

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "UserMap"

# Headers
ws["A1"] = "Name"
ws["B1"] = "Email"

# Team member data
team_members = [
    ("Chris Greer", "chris.greer@flex.com"),
    ("Mike Keane", "mike.keane@flex.com"),
    ("Tess York", "tess.york@flex.com"),
    ("Shawna Wass", "shawna.wass@flex.com"),
    ("Mike Sun", "mike.sun@flex.com"),
    ("Kim Jordan", "kim.jordan@flex.com"),
    ("Tia Waldron", "tia.waldron@flex.com"),
]

for i, (name, email) in enumerate(team_members, start=2):
    ws[f"A{i}"] = name
    ws[f"B{i}"] = email

# Format as a named table
table = Table(displayName="UserMap", ref="A1:B8")
style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
table.tableStyleInfo = style
ws.add_table(table)

# Auto-fit column widths (approximate)
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 28

output_path = "UserMapping.xlsx"
wb.save(output_path)
print(f"Created {output_path} with {len(team_members)} team members in table 'UserMap'")
